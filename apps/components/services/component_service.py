import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from apps.components.models import Component

class ComponentService:
    @staticmethod
    def export_project_components_to_excel(project):
        """
        Exports all components of a project into an Excel (.xlsx) spreadsheet.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Linh kiện Dự án"

        # Styles
        title_font = Font(name="Arial", size=14, bold=True, color="15803D")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        data_font = Font(name="Arial", size=10)
        total_font = Font(name="Arial", size=10, bold=True, color="15803D")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # Title Block
        ws.merge_cells('A1:F1')
        ws['A1'] = f"DANH SÁCH LINH KIỆN - DỰ ÁN: {project.name.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Header Row
        headers = ["STT", "Tên Linh Kiện", "Số Lượng", "Đơn Giá (VNĐ)", "Thành Tiền (VNĐ)", "Shop / Cửa Hàng", "Ghi Chú"]
        ws.append([]) # Blank row 2
        
        ws.cell(row=3, column=1, value="STT").font = header_font
        ws.cell(row=3, column=2, value="Tên Linh Kiện").font = header_font
        ws.cell(row=3, column=3, value="Số Lượng").font = header_font
        ws.cell(row=3, column=4, value="Đơn Giá (VNĐ)").font = header_font
        ws.cell(row=3, column=5, value="Thành Tiền (VNĐ)").font = header_font
        ws.cell(row=3, column=6, value="Shop / Nơi Mua").font = header_font
        ws.cell(row=3, column=7, value="Ghi Chú").font = header_font

        for col in range(1, 8):
            ws.cell(row=3, column=col).fill = header_fill
            ws.cell(row=3, column=col).alignment = Alignment(horizontal='center', vertical='center')

        components = Component.objects.filter(project=project).order_by('created_at')
        current_row = 4
        total_sum = 0

        for idx, item in enumerate(components, start=1):
            ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=2, value=item.name)
            ws.cell(row=current_row, column=3, value=item.quantity).alignment = Alignment(horizontal='right')
            ws.cell(row=current_row, column=4, value=float(item.unit_price)).number_format = '#,##0'
            ws.cell(row=current_row, column=5, value=float(item.total_price)).number_format = '#,##0'
            ws.cell(row=current_row, column=6, value=item.shop or "")
            ws.cell(row=current_row, column=7, value=item.notes or "")

            for col in range(1, 8):
                ws.cell(row=current_row, column=col).font = data_font
                ws.cell(row=current_row, column=col).border = thin_border

            total_sum += item.total_price
            current_row += 1

        # Summary Row
        ws.cell(row=current_row, column=1, value="TỔNG CỘNG").font = total_font
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=5, value=float(total_sum)).font = total_font
        ws.cell(row=current_row, column=5).number_format = '#,##0'

        # Auto adjust column widths
        column_widths = {'A': 8, 'B': 30, 'C': 12, 'D': 18, 'E': 20, 'F': 25, 'G': 30}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # HTTP Response
        filename = f"LinhKien_{project.name.replace(' ', '_')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
